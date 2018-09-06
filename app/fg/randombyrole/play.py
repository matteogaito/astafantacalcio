from app import app,db
from flask import render_template,session,redirect,url_for,request, send_file
from openpyxl import load_workbook, Workbook
import os
import requests
import random
from sqlalchemy import create_engine, MetaData, Table
from sqlalchemy import Column, Integer, String
from sqlalchemy.sql import select, update
from sqlalchemy.sql.expression import func

from app.models import Teams
#from app.models import Giocatori, Portieri, Centrocampisti, Difensori, Attaccanti


class Estrazione():
    def __init__(self, tipo):
        self.tipo = tipo
        if 'path_xls' in session:
            self.path_lista_xls = session['path_xls'] + "/lista.xlsx"
            self.path_lega = session['path_xls']
            self.lega_name = session['lega_name']
            self.legadb = self._inizializeDBLega()
            self.con_legadb = self.legadb.connect()
        else:
            app.logger.error("Error, path_xls non e' in sessione")
            raise "PathXLSEmpty"

    def _error(which):
        print(which)

    def _inizializeDBLega(self):
        legadb =  create_engine('sqlite:///' + self.lega_name + '.db', echo=True)
        return legadb

    def _getRoleModel(self, role):
        metadata = MetaData(bind=self.legadb)
        players = Table(role, metadata,
            Column('id', Integer, primary_key=True),
            Column('name', String(50)),
            Column('player_row', Integer),
            Column('club', String(50)),
            Column('quotazione', Integer),
            Column('stato', Integer)
                        )

        return players, metadata


    def _addPlayertoDB(self, role, player_info):
        app.logger.info("aggiungo {} sulla tabella {}".format(role, player_info['name']))

        players, metadata = self._getRoleModel(role)

        try:
            players.create(self.legadb)
        except:
            app.logger.info("Table already created")

        self.con_legadb.execute(players.insert(), name = player_info['name'],
                                  player_row = player_info['player_row'],
                                  club = player_info['club'],
                                  quotazione = player_info['quotazione'],
                                  stato = 0
                )

    def _createPlayersList(self):
        players_source_list = self.path_lista_xls
        wb = load_workbook(players_source_list)
        ws_role = wb[self.role]
        offset = 2
        players_list = []
        for row in ws_role.rows:
            if row[0].row <= offset:
                continue
            if len(row) == 8:
                if row[7].value == "aggiudicato":
                    continue
            player_info = {}
            player_info['name'] = row[2].value
            player_info['player_row'] = row[0].row
            player_info['club'] = row[3].value
            player_info['quotazione'] = row[4].value
            app.logger.info("added {} to player database".format(player_info['name']))
            self._addPlayertoDB(self.role, player_info)
            players_list.append(player_info)

        return players_list

    def _populateRole(self, role):
        if not self.legadb.dialect.has_table(self.legadb, role):
            app.logger.info("Tabella {} non presente".format(role))
            self._createPlayersList()

    def _removePlayerFromList(self, role, index, stato):
        players, metadata = self._getRoleModel(role)
        stmt = update(players).where(players.c.id==5).values(stato=stato)
        self.con_legadb.execute(stmt)

    def incorso(self, *positional_parameters, **keyword_parameters):
        if ('quale' in keyword_parameters):
            session.modified = True
            quale = keyword_parameters['quale']
            session['estrazione_in_corso'] = url_for('estrai', players=quale)
            app.logger.info("Sovrascrivo estrazione in corso")
        else:
            app.logger.info("Nessun parametro quale")
            try:
                return session['estrazione_in_corso']
            except:
                return self._error(which="Missing parameter in function 'incorso'")

    def conver_row_to_player_info(self, row):
        player_info = {}
        player_id = 0
        player_info['name'] = row.name
        player_info['player_row'] = row.player_row
        player_info['club'] = row.club
        player_info['quotazione'] = row.quotazione
        player_info['stato'] = row.stato
        player_info['img_url'] = self._getPlayerImage(row.name)
        player_id = row.id

        return player_info, player_id

    def getRandomPlayer(self, role):
        self.role = role
        self._populateRole(role)
        players, metadata = self._getRoleModel(role)
        s = select([players]).where(players.c.stato == 0).order_by(func.random()).limit(1)
        result = self.con_legadb.execute(s)
        for row in result:
            player = row
            print(player.name)

        player_info, player_id = self.conver_row_to_player_info(player)
        return player_info, player_id

    def getPlayerbyIndex(self, role, index):
        self.role = role
        self._populateRole(role)
        players, metadata = self._getRoleModel(role)
        s = select([players]).where(players.c.id == index)
        result = self.con_legadb.execute(s)

        for row in result:
            player = row
            print(player.name)

        player_info, player_id = self.conver_row_to_player_info(player)
        return player_info, player_id

    def _campioncinoImageChecker(self, nameurl):
        campioncino_url = "https://content.fantagazzetta.com/web/campioncini/card/"
        complete_url = campioncino_url + nameurl + '.jpg'
        session = requests.Session()
        richiesta = session.get(complete_url, verify=True, headers=app.config['HTTP_HEADERS'])
        if (richiesta.status_code == 200):
            return complete_url
        else:
            return False

    def _getPlayerImage(self, name):
        app.logger.info("Download Url Campioncino")
        nameurls = []
        list_exception = ["ROSSI F"]
        if name not in list_exception:
            # Primo: al posto degli spazi metto dash
            nameurls.append(name.replace(" ", "-"))
            # Secondo: prendo solo la prima parola
            nameurls.append(name.split()[0])
            # Default image
            nameurls.append('NO-CAMPIONCINO')
        else:
            nameurls.append('NO-CAMPIONCINO')
        for nameurl in nameurls:
            imageurl = self._campioncinoImageChecker(nameurl)
            if imageurl is not False:
                app.logger.info("{} immagine {}".format(name, imageurl))
                return imageurl

    def _addPlayerImage(self, index):
        session[self.players_list_name][index]['img_url'] = self._getPlayerImage(session[self.players_list_name][index]['name'])

    def scarta(self, role, index):
        self._removePlayerFromList(role, index, 1)

    def getTeamFile(self, team):
        return self.path_lega + "/" + team.replace(" ", "_").replace("'", "") + ".xlsx"

    def assignToTeam(self, team, player, index, players, cost):
        team_xls = self.getTeamFile(team)
        if not os.path.exists(team_xls):
            app.logger.info("File {} non esistente, lo creo".format(team_xls))
            wb = Workbook()
            wb.save(team_xls)

        wb = load_workbook(team_xls)
        sheet = 'rosa'
        if sheet not in wb.sheetnames:
            app.logger.info("Sheet {} non esistente, lo creo".format(sheet))
            wb.create_sheet(sheet)
            wb.save(team_xls)

        for s in wb.sheetnames:
            print(s)
            if s == "Sheet":
                wb.remove(wb.get_sheet_by_name(s))
                wb.save(team_xls)

        # Salva su file di team
        wb = load_workbook(team_xls)
        ws_role = wb[sheet]

        ws_rows = list(ws_role.rows)
        ws_last_row_index = len(ws_rows) -1

        if ws_rows[ws_last_row_index] is not None:
            position = ws_last_row_index + 2

        # Salvo su file team
        ws_role.cell(column=1, row=position, value=player['name'])
        ws_role.cell(column=2, row=position, value=cost)
        wb.save(team_xls)

        # Salvo su file di lega
        players_source_list = self.path_lista_xls
        wb = load_workbook(players_source_list)
        ws_role = wb[players]
        position = player['player_row']
        ws_role.cell(column=8, row=position, value = "aggiudicato")
        wb.save(players_source_list)
        # Remove from list
        self._removePlayerFromList(players, index, 2)

@app.route('/fg/randombyrole/play')
def play():
    teams = Teams.query.filter_by(leghe_id=session['lega_id'])

    return render_template('fg-randombyrole.html', teams = teams)

@app.route('/fg/randombyrole/estrai')
def estrai():
    try:
        players = request.args['players']
    except:
        return redirect(url_for('error', missing_role=True))

    Estrazione(tipo='randombyrole').incorso(quale=players)
    app.logger.info("Estrazione {}".format(players))

    player, index = Estrazione(tipo='randombyrole').getRandomPlayer(players)
    app.logger.info("Estratto {} indice".format(player['name'], index))

    return render_template('fg-randombyrole-play.html', player = player, num=index, action = "estrai")

@app.route('/fg/randombyrole/assegna')
def assegna():
    try:
        players = request.args['players']
    except:
        return redirect(url_for('error', missing_param="players"))
    try:
        index = int(request.args['index'])
    except:
        return redirect(url_for('error', missing_param="player_id"))

    teams = Teams.query.filter_by(leghe_id=session['lega_id'])

    player, index = Estrazione(tipo='randombyrole').getPlayerbyIndex(players, index)

    return render_template('fg-randombyrole-play.html', player = player, num=index, teams = teams, action = "assegna")

@app.route('/fg/randombyrole/scarta')
def scarta():
    try:
        players = request.args['players']
    except:
        return redirect(url_for('error', missing_param="role"))

    try:
        index = int(request.args['index'])
    except:
        return redirect(url_for('error', missing_param="player_id"))

    Estrazione(tipo='randombyrole').scarta(players, index)

    return redirect(Estrazione(tipo='randombyrole').incorso())

@app.route('/fg/randombyrole/downloadlista')
def downloadlist():
    try:
        teamid = request.args['teamid']
    except:
        return redirect(url_for('error', missing_role="teamid"))

    team = Teams.query.filter_by( id = teamid )
    team_list = Estrazione(tipo='randombyrole').getTeamFile(team[0].name)
    filename = team_list.split('/')[-1]
    print(filename)

    return send_file(team_list, as_attachment=True)

@app.route('/fg/randombyrole/confermato', methods=["POST"])
def confermato():
    session.modified = True

    try:
        players = request.form['players']
    except:
        return redirect(url_for('error', missing_param="role"))

    try:
        index = int(request.form['index'])
    except:
        return redirect(url_for('error', missing_param="player_id"))

    try:
        cost = int(request.form['cost'])
    except:
        return redirect(url_for('error', missing_param="cost"))

    player, index = Estrazione(tipo='randombyrole').getPlayerbyIndex(players, index)

    if request.form['action'] == "assegna":
        team = request.form['team[1][]']
        Estrazione(tipo='randombyrole').assignToTeam(team, player, index, players, cost)

    if request.form['action'] == "scarta":
        Estrazione(tipo='randombyrole').scarta(players, index)

    return redirect(Estrazione(tipo='randombyrole').incorso())
